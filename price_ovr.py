# -*- coding: utf-8 -*-
import sys
import time

from openpyxl import load_workbook
from datetime import datetime
from gui import Ui_Converor
from PyQt5.QtWidgets import QApplication
from PyQt5 import QtWidgets
from functools import lru_cache




class Price_convertor(QtWidgets.QMainWindow):
    def __init__(self):
        self.ui = Ui_Converor()
        super().__init__()
        self.ui.setupUi(self)
        self.l1 = """"""
        res = []
        self.file_name = ""
        self.vinList = list()
        self.priceList = list()
        self.MultiLineString = """"""
        self.ui.Open_file.clicked.connect(lambda x: self.showDialog())
        self.ui.Run.clicked.connect(lambda x: self.start_work())

    def now(self):
        now = datetime.now()
        self.current_time = now.strftime("%H:%M:%S")

    def showDialog(self):
        fname = QtWidgets.QFileDialog.getOpenFileName(self, 'Open file', '*.xlsx')[0]
        name_index_ = fname.rfind("/")
        self.file_name = fname
        self.ui.incoming_file.setText(fname[name_index_ + 1:])
        self.now()
        self.ui.start_at.setText(f'Початок: {datetime.now().strftime("%H:%M:%S")}')
        self.start = datetime.now().strftime("%H:%M:%S")

        print(f'start at {self.start}')

    def f1(self):
        wb = load_workbook(self.file_name, data_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.iter_rows(min_row=2, ):
            yield [cell.value for cell in row if cell.value]

    def split_file(self, suffix, l1):

        with open(f"res{suffix}.csv", "w") as file:
            file.write(l1)

    def start_work(self):
        # print(list(self.f1()))

        l1 = """"""
        res = []
        res_f1 = list(self.f1())
        for i in res_f1:
            res += i

        vin = 0
        price = 1
        length = int(len(res))

        for id, item in enumerate(res, 1):
            if price > length:
                break
            if id % 500000 == 0:
                print(id)
                self.split_file(id, l1)
                l1 = """"""
            if isinstance(res[price], (float, int)):
                l1 += ";".join([res[vin], res[vin], res[vin], "5", "0", f"{res[price]}", "\n"])

                vin += 2
                price += 2

            else:
                vin += 1
                price += 1

        self.split_file("_last", l1)

        self.end = datetime.now().strftime("%H:%M:%S")

        print(f'End at {self.end }')
        self.end = datetime.now().strftime("%H:%M:%S")
        self.ui.end_at.setText(f"Закінчили: {datetime.now().strftime('%H:%M:%S')}")
        self.ui.end_at.setStyleSheet("color: green")


def main():
    app = QApplication(sys.argv)
    w = Price_convertor()
    w.show()
    app.exec_()


if __name__ == '__main__':
    main()

